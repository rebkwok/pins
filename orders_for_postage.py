"""
First fix up xlsx file for shipping address so that city, county, postcode 
are the last items. County can be an empty line.
"""
import csv
import openpyxl
import sys

from pathlib import Path

def main(xlsx_file):
    xlsx_path = Path(xlsx_file)
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active

    headers = [
        col[0].value for col in sheet.iter_cols(1, sheet.max_column)
    ]

    filter_to_headers = ["Name", "Email", "Shipping address", "Submission date", "Reference", "Total items"]
    
    header_indicies = {
        headers.index(header): header for header in filter_to_headers
    }
    
    data = []

    for row in range(1, sheet.max_row):
        row_data = {}
        vals = [col[row].value for col in sheet.iter_cols(1, sheet.max_column)]
        if "Becky Smith" in vals:
            continue
        if vals[headers.index("Total items")] == 0:
            continue
        for i, val in enumerate(vals):
            if i in header_indicies:
                if header_indicies[i] == "Shipping address":
                    address = val.strip().split("\n")
                    if len(address) == 1:
                        address = val.strip().split(",")
                        address = [v.strip() for v in address]
                    
                    if len(address) == 1:
                        split_address = val.strip().split(" ")
                        address = split_address[:-2]
                        address.append(" ".join(split_address[-2:]))

                    postcode = address[-1]
                    county = address[-2]
                    city = address[-3]
                    address_lines = address[0:-3]
                    if len(address_lines) > 3:
                        last = ", ".join(address_lines[2:])
                        address_lines = [*address_lines[:2], last]
                    for i in range(3):
                        if len(address_lines) > i:
                            line = address_lines[i]
                        else:
                            line = ""
                        row_data[f"Address line {i + 1}"] = line

                    row_data["City"] = city
                    row_data["County"] = county
                    row_data["Postcode"] = postcode
                else:
                    if header_indicies[i] == "Submission date":
                        val = val.strftime("%Y%m%d")
                    if header_indicies[i] == "Total items":
                        row_data["Weight"] = val * 0.7
                    row_data[header_indicies[i]] = val
        data.append(row_data)
    
    outfile = Path(f"{xlsx_path.stem}_for_postage.csv")
    with outfile.open("wt") as out:
        fieldnames = data[0].keys()
        writer = csv.DictWriter(out, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


if __name__ == "__main__":
    main(sys.argv[1])
