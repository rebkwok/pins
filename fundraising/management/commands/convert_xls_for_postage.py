"""
First fix up xlsx file for shipping address so that city, [county], postcode 
are the last items. County is optional.
"""
import csv
import openpyxl
import re

from pathlib import Path

from django.core.management.base import BaseCommand



POSTCODE_RE = re.compile('[A-Z]{1,2}[0-9]{1,2}[A-Z]? ?[0-9][A-Z]{2}', flags=re.I)


def find_postcode(address):
    """Return a normalised postcode if valid, or None if not."""
    found = POSTCODE_RE.findall(address)
    if found:
        return found[0]
    return None


class Command(BaseCommand):
    help = 'Convert xlsx file to csv for click & drop upload'

    def add_arguments(self, parser):
        parser.add_argument(
            "xlsx_path",
            type=Path,
            help="Path to xlsx file downloaded from order form admin page",
        )
        return super().add_arguments(parser)

    def handle(self, xlsx_path, **kwargs):
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active

        headers = [
            col[0].value for col in sheet.iter_cols(1, sheet.max_column)
        ]

        filter_to_headers = ["Name", "Email", "Shipping address", "Submission date", "Reference", "Total items"]
        
        header_indicies = {
            headers.index(header): header for header in filter_to_headers
        }
        
        data = []

        for row_num, row in enumerate(range(1, sheet.max_row), start=1):
            row_data = {}
            vals = [col[row].value for col in sheet.iter_cols(1, sheet.max_column)]
            if "Becky Smith" in vals:
                continue
            if not vals[headers.index("Submission date")]:
                continue
            if vals[headers.index("Total items")] == 0:
                continue
            for i, val in enumerate(vals):
                if i in header_indicies:
                    if header_indicies[i] == "Shipping address":
                        
                        # First find postcode
                        postcode = find_postcode(val)
                        if postcode:
                            val = val.replace(postcode, "")
                            postcode = postcode.upper()

                        # strip any possible trailing commas or spaces and split by new lines
                        address = val.split("\n")
                        address = [line.strip(",").strip() for line in address if line]

                        # if we only have one line, it wasn't entered as a multiline address, 
                        # try splitting by commas
                        if len(address) == 1:
                            address = val.strip().split(",")
                            address = [v.strip() for v in address]
                        
                        # still one line, it was entered as a single line with no delimiter
                        # raise exception, can't determine address from here
                        if len(address) == 1:
                            raise ValueError(f"Can't parse address on row {row_num}: {val}")

                        if len(address) == 2:
                            address_lines = [address[0]]
                            city = address[1]
                            county = ""
                        else:
                            county = address[-1]
                            city = address[-2]
                            address_lines = address[0:-2]
                        
                        for i in range(3):
                            if len(address_lines) > i:
                                if i == 2:
                                    line = ", ".join(address_lines[i:])
                                else:
                                    line = address_lines[i]
                            else:
                                line = ""
                            row_data[f"Address line {i + 1}"] = line

                        row_data["City"] = city
                        row_data["County"] = county
                        row_data["Postcode"] = postcode
                    else:
                        if header_indicies[i] == "Submission date":
                            val = val.strftime("%Y%m%d")
                        if header_indicies[i] == "Total items":
                            weights = {
                                1: 0.8,
                                2: 1.3,
                                3: 1.9,
                                4: 2.5,
                                5: 3.1,
                                20: 12,
                            }
                            row_data["Weight"] = weights[val]
                        row_data[header_indicies[i]] = val
            data.append(row_data)
        
        outfile = Path(f"{xlsx_path.stem}_for_postage.csv")
        with outfile.open("wt") as out:
            fieldnames = data[0].keys()
            writer = csv.DictWriter(out, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)